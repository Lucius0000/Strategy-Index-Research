# -*- coding: utf-8 -*-
"""
各牛熊周期下 ETF/指数 表现 & 可视化
- 标的：['510300.SS','510310.SS','510330.SS','000300.SS']
- 原始数据：output/raw_data/<ticker>.xlsx
- 摘要：    output/raw_data/CSI300_brief.txt
- 指标：    output/ETF_metrics.xlsx（合并一个文件，中文列名）
"""

from __future__ import annotations
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import warnings
warnings.filterwarnings("ignore")

import time
from datetime import datetime
import pandas as pd
import numpy as np
import yfinance as yf
import matplotlib
import matplotlib.pyplot as plt
import mplfinance as mpf
import os
import re

# 设置代理
os.environ['http_proxy'] = 'http://127.0.0.1:7890'
os.environ['https_proxy'] = 'http://127.0.0.1:7890'

# ---------------- 配置区 ----------------
# 指数清单（一般只定义 1 个，且作为基准）
INDEX_TICKERS: List[str] = ['^GSPC']
# ETF 清单
ETF_TICKERS:   List[str] = ['SPY','VOO']
TICKERS: List[str] = INDEX_TICKERS + ETF_TICKERS

# 周期来源 ，manual 是配置区指定list，table 是从本地表格解析
PERIOD_SOURCE: str = 'table'
# 市场周期表格路径
PERIODS_TABLE_PATH: str = 'data/SP500.xlsx'
# 指定 sheet
PERIODS_TABLE_SHEET: Optional[str | int] = None

# 当使用 manual 时，请为每个元素补充“周期类型”，形如 ('YYYY-MM','YYYY-MM','牛市')
PERIODS: list[tuple[str, str, str]] = [
    ('2005-06','2007-10','牛市'),
    ('2007-10','2008-10','熊市'),
    ('2008-11','2009-08','牛市'),
    ('2009-08','2014-07','震荡'),
    ('2014-03','2015-06','牛市'),
    ('2015-06','2016-01','熊市'),
    ('2016-02','2018-01','牛市'),
    ('2018-02','2019-01','熊市'),
    ('2019-01','2021-02','牛市'),
    ('2021-02','2025-10','熊市'),
]

# 是否需要输出可视化图表
VISUALIZE: bool = False

RISK_FREE_ANNUAL = 0.027
ETF_ADJUST = ""

DIR_RAW = Path("output/raw_data"); DIR_RAW.mkdir(parents=True, exist_ok=True)
DIR_OUT = Path("output");          DIR_OUT.mkdir(parents=True, exist_ok=True)

BRIEF_TXT = DIR_RAW / "SP500_brief.txt"

# 表格名称
ALL_METRICS_XLSX = DIR_OUT / f"{INDEX_TICKERS[0]}_ETF.xlsx"

TE_LOOKBACK = 252
MAX_RETRY = 2
RETRY_SLEEP_SECONDS = 1
# --------------------------------------

# ---------- 中文字体 ----------
def setup_chinese_font():
    matplotlib.rcParams['axes.unicode_minus'] = False
    matplotlib.rcParams['font.sans-serif'] = ['SimHei','Microsoft YaHei','Arial Unicode MS','Noto Sans CJK SC']
setup_chinese_font()

# ---------- 周期读取：从表格生成 PERIODS ----------
def _normalize_periods_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    将常见中文列名归一为 ['start','end','type','index'] 便于后续处理。
    仅 start/end 必需；type/index 可选。
    """
    col_map = {
        '开始时间': 'start', '开始日期': 'start', '开始': 'start', '起始时间': 'start', '起始': 'start',
        '结束时间': 'end',   '结束日期': 'end',   '结束': 'end',   '终止时间': 'end',   '终止': 'end',
        '周期类型': 'type',  '类型': 'type', '牛熊': 'type',
        '指数': 'index', '标的': 'index', '代码': 'index', '指数代码': 'index',
    }
    ren = {c: col_map.get(str(c).strip(), c) for c in df.columns}
    df = df.rename(columns=ren)
    return df

def _to_yyyymm(x) -> Optional[str]:
    if pd.isna(x):
        return None
    try:
        dt = pd.to_datetime(x, errors='coerce')
        if pd.isna(dt):
            return None
        return dt.strftime('%Y-%m')
    except Exception:
        return None

def load_periods_from_table(path: str | Path,
                            sheet: Optional[str | int] = None,
                            index_filter: Optional[str] = None) -> list[tuple[str, str, str]]:
    """
    从 data 下的表（xlsx/xls/csv）读取牛熊区间，输出为 [('YYYY-MM','YYYY-MM','类型'), ...]。
    - 若存在 'index' 列且传入 index_filter（如 '000300' 或 '000300.SS'），则据此筛选。
    - 自动剔除 start/end 缺失或无效的行。
    - 去重并按开始时间排序。
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"周期表不存在：{path}")

    # 读取
    if path.suffix.lower() in {'.xlsx', '.xls'}:
        df = pd.read_excel(path, sheet_name=sheet)
    elif path.suffix.lower() == '.csv':
        df = pd.read_csv(path, encoding='utf-8')
    else:
        raise ValueError(f"不支持的文件类型：{path.suffix}")

    if isinstance(df, dict):  # 指定表名返回的是 dict
        # 若 sheet=None 则取第一个
        df = next(iter(df.values()))

    df = _normalize_periods_columns(df)

    # 可选：按指数过滤（若表里放了多只指数的区间）
    if 'index' in df.columns and index_filter:
        key = index_filter.replace('.SS','').replace('.SZ','')
        df = df.loc[df['index'].astype(str).str.contains(key, na=False)]

    # 提取并规范化
    starts = df.get('start', pd.Series(dtype=object)).apply(_to_yyyymm)
    ends   = df.get('end',   pd.Series(dtype=object)).apply(_to_yyyymm)
    types  = df.get('type',  pd.Series(dtype=object)).fillna('')

    out: list[tuple[str, str, str]] = []
    for s, e, t in zip(starts, ends, types):
        if not s or not e:
            continue
        # 过滤非法顺序
        try:
            if pd.to_datetime(s) > pd.to_datetime(e):
                continue
        except Exception:
            continue
        out.append((s, e, str(t)))

    # 去重/排序
    out = sorted(set(out), key=lambda x: (x[0], x[1], x[2]))
    return out

# 将 manual/table 两种来源统一为 3 元组列表
def normalize_periods(period_source: str,
                      periods_cfg: list[tuple[str, str, str]],
                      table_path: str,
                      table_sheet: Optional[str | int]) -> list[tuple[str, str, str]]:
    if period_source.lower() == 'table':
        return load_periods_from_table(table_path, sheet=table_sheet)
    # manual：直接使用配置区 PERIODS（要求为三元组）
    out = []
    for it in periods_cfg:
        if len(it) != 3:
            raise ValueError("当 PERIOD_SOURCE='manual' 时，PERIODS 中每个元素需形如 ('YYYY-MM','YYYY-MM','牛市/熊市/震荡')")
        out.append(it)
    return out

# ---------- 下载 & 预处理 ----------
def month_start(s: str) -> pd.Timestamp:
    # 'YYYY-MM' -> 当月月初 00:00:00
    return pd.to_datetime(s, format='%Y-%m').to_period('M').start_time

def month_end(s: str) -> pd.Timestamp:
    # 'YYYY-MM' -> 当月月末 23:59:59.999999999
    return pd.to_datetime(s, format='%Y-%m').to_period('M').end_time

# （修改）指数判定改为基于配置清单
def is_index_ticker(ticker: str) -> bool:
    return ticker in set(INDEX_TICKERS)

# -------------------- 关键改动：统一使用 yfinance 拉取 --------------------
START_DATE_FALLBACK = "2000-01-01"
END_DATE_TODAY = datetime.now().strftime("%Y-%m-%d")

def _normalize_price_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    mapping = {'date':'Date','日期':'Date','时间':'Date',
               'open':'Open','开盘':'Open',
               'high':'High','最高':'High',
               'low':'Low','最低':'Low',
               'close':'Close','收盘':'Close','收盘价':'Close',
               'volume':'Volume','成交量':'Volume'}
    df = df.rename(columns={c: mapping.get(c, c) for c in df.columns})
    keep = [c for c in ['Date','Open','High','Low','Close','Volume'] if c in df.columns]
    df = df.loc[:, keep].copy()
    if 'Date' in df.columns:
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        df = df.dropna(subset=['Date']).drop_duplicates(subset=['Date']).sort_values('Date').set_index('Date')
    return df

def yf_fetch_full_history(ticker: str) -> pd.DataFrame:
    """
    使用 yfinance 拉取日频 OHLCV，未复权（auto_adjust=False），
    兼容美股/港股/沪深（如 .SS/.SZ）/指数（^GSPC、^HSI 等）。
    """
    last_err = None
    for _ in range(MAX_RETRY):
        try:
            tk = yf.Ticker(ticker)
            hist = tk.history(start=START_DATE_FALLBACK, end=END_DATE_TODAY,
                              interval="1d", auto_adjust=False)
            if hist is None or hist.empty:
                # 有些指数需要用 yf.download 批量接口更稳
                hist = yf.download(ticker, start=START_DATE_FALLBACK, end=END_DATE_TODAY,
                                   interval="1d", auto_adjust=False, progress=False)
            if hist is not None and not hist.empty:
                out = hist.copy()

                # ======= 修改：替换省略号，保证可运行并且列名统一 =======
                keep_cols = []
                for c in ["Open", "High", "Low", "Close", "Volume"]:
                    if c in out.columns:
                        keep_cols.append(c)
                if not keep_cols:
                    return pd.DataFrame()

                out = out.loc[:, keep_cols].copy()
                out.index = pd.to_datetime(out.index)

                # ↓↓↓ 新增：统一去时区，避免下游写 Excel 报错
                if isinstance(out.index, pd.DatetimeIndex) and out.index.tz is not None:
                    out.index = out.index.tz_convert('UTC').tz_localize(None)
                out = out[~out.index.duplicated(keep='last')].sort_index()

                out = out.rename_axis('Date').reset_index()
                return _normalize_price_columns(out)
                # ======= 修改结束 =======
        except Exception as e:
            last_err = e
            time.sleep(RETRY_SLEEP_SECONDS)
    raise RuntimeError(f"{ticker} 下载失败：{last_err}")

def fetch_full_history_by_ticker(ticker: str) -> pd.DataFrame:
    return yf_fetch_full_history(ticker)

def fetch_etf_dividend_cumulative(ticker: str) -> pd.Series:
    """
    使用 yfinance 的现金分红序列（每次派息金额），
    转为“累计分红”Series（升序，索引为日期）。
    对于指数（^GSPC、^HSI 等）或无分红标的，返回空 Series。
    """
    if is_index_ticker(ticker):
        return pd.Series(dtype=float)
    last_err = None
    for _ in range(MAX_RETRY):
        try:
            tk = yf.Ticker(ticker)
            div = tk.dividends  # pandas Series: index=date, values=amount
            if div is None or len(div) == 0:
                return pd.Series(dtype=float)

            div = div.copy()
            div.index = pd.to_datetime(div.index)

            # ==== 新增：统一去掉时区并规范为“日期索引” ====
            if isinstance(div.index, pd.DatetimeIndex) and div.index.tz is not None:
                # 不关心具体时刻，只要交易“日期”，因此先转UTC再去时区并规范到日期
                div.index = div.index.tz_convert('UTC').tz_localize(None)
            # 规范成纯日期（00:00:00），避免和 month_end 比较出错
            div.index = div.index.normalize()
            # ===========================================

            div = div.sort_index()
            div = div.groupby(div.index).sum()  # 同日多条求和
            cum = div.cumsum()
            cum.name = "CumDiv"
            cum = cum[~cum.index.duplicated(keep='last')]
            return cum
        except Exception as e:
            last_err = e
            time.sleep(RETRY_SLEEP_SECONDS)
    raise RuntimeError(f"{ticker} 分红下载失败：{last_err}")

# ------------------------------------------------------------------

def save_raw_to_excel(ticker: str, df: pd.DataFrame) -> Path:
    path = DIR_RAW / f"{ticker}.xlsx"
    out = df.copy()
    # 确保索引无时区
    if isinstance(out.index, pd.DatetimeIndex) and out.index.tz is not None:
        out.index = out.index.tz_convert('UTC').tz_localize(None)
    out.index.name = "Date"
    out.to_excel(path, sheet_name="Price", engine="openpyxl")
    return path

# ---------- 指标（年化） ----------
def annualize_return(daily_ret: pd.Series) -> float:
    r = daily_ret.dropna()
    if len(r)==0: return np.nan
    return ((1+r).prod() ** (252.0/len(r)) - 1.0)

def annualize_vol(daily_ret: pd.Series) -> float:
    r = daily_ret.dropna()
    if len(r)<2: return np.nan
    return r.std(ddof=1)*np.sqrt(252.0)

def max_drawdown(nav: pd.Series) -> float:
    x = nav.dropna()
    if x.empty: return np.nan
    return (x/x.cummax()-1).min()

def sharpe_ratio(daily_ret: pd.Series, rf_annual: float=0.0) -> float:
    r = daily_ret.dropna()
    if len(r)<2: return np.nan
    rf_daily = (1+rf_annual)**(1/252.0)-1.0
    ex = r - rf_daily
    mu = ex.mean()*252.0
    sigma = ex.std(ddof=1)*np.sqrt(252.0)
    return np.nan if sigma==0 or np.isnan(sigma) else mu/sigma

def information_ratio(daily_ret: pd.Series, bench_daily_ret: pd.Series) -> float:
    r = pd.concat([daily_ret, bench_daily_ret], axis=1, keys=['etf','bench']).dropna()
    if len(r)<2: return np.nan
    active = r['etf'] - r['bench']
    alpha = active.mean()*252.0
    te = active.std(ddof=1)*np.sqrt(252.0)
    return np.nan if te==0 or np.isnan(te) else alpha/te

def tracking_error_1y(etf_close: pd.Series, idx_close: pd.Series) -> float:
    etf = etf_close.dropna().iloc[-TE_LOOKBACK:]
    idx = idx_close.dropna().iloc[-TE_LOOKBACK:]
    df = pd.concat([etf, idx], axis=1, keys=['etf','idx']).dropna()
    if len(df)<30: return np.nan
    active = df['etf'].pct_change() - df['idx'].pct_change()
    active = active.dropna()
    if len(active)<2: return np.nan
    return active.std(ddof=1)*np.sqrt(252.0)

# 区间年化跟踪误差：在任意时间段内用日度主动收益的波动率 * sqrt(252)
def tracking_error_annualized_period(etf_close: pd.Series, idx_close: pd.Series) -> float:
    df = pd.concat([etf_close, idx_close], axis=1, keys=['etf','idx']).dropna()
    if len(df) < 30:
        return np.nan
    active = df['etf'].pct_change() - df['idx'].pct_change()
    active = active.dropna()
    if len(active) < 2:
        return np.nan
    return active.std(ddof=1) * np.sqrt(252.0)

# 区间分红 / 年化分红率
def dividend_in_period(cum_div: pd.Series, start_dt: pd.Timestamp, end_dt: pd.Timestamp) -> float:
    """
    累计分红序列 -> 区间分红：end_asof - start_before_asof
    - start_before_asof: 严格小于 start 的最后一个累计分红（区间前的累计值）
    - end_asof: 小于等于 end 的最后一个累计分红
    """
    if cum_div is None or cum_div.empty:
        return np.nan
    cum_div = cum_div.sort_index()
    # 截止 end 的累计值（asof）
    end_slice = cum_div.loc[cum_div.index <= end_dt]
    end_val = float(end_slice.iloc[-1]) if len(end_slice) else 0.0
    # start 之前的累计值（strictly < start）
    start_slice = cum_div.loc[cum_div.index < start_dt]
    start_val = float(start_slice.iloc[-1]) if len(start_slice) else 0.0
    div = end_val - start_val
    # 容错：若出现负值（数据瑕疵），归零
    return div if div >= 0 else 0.0

def dividend_yield_annualized(div_amount: float, seg_close: pd.Series) -> float:
    """
    年化分红率 = 区间分红 / 区间平均收盘价 * (252 / 交易日数)
    返回数值（例如 0.03 表示 3%）
    """
    seg_close = seg_close.dropna()
    if len(seg_close) == 0 or pd.isna(div_amount):
        return np.nan
    avg_price = float(seg_close.mean())
    if avg_price <= 0:
        return np.nan
    ann_factor = 252.0 / len(seg_close)
    return (div_amount / avg_price) * ann_factor

# ======= 新增：与沪深300一致的“按牛/熊/震荡分组汇总”算法（基于原始日度数据） =======
def normalize_market_cycle(raw: str) -> str:
    """
    将周期类型归一为三类：牛市、熊市、震荡
    说明：
    - 震荡类：只要包含“震荡”二字即可（例如：震荡修复、震荡调整）
    - 其余：包含“牛” -> 牛市；包含“熊” -> 熊市
    """
    t = "" if raw is None else str(raw)
    if re.search(r"震荡", t):
        return "震荡"
    if re.search(r"牛", t):
        return "牛市"
    if re.search(r"熊", t):
        return "熊市"
    return t.strip()

def _active_ret_from_close(etf_close: pd.Series, bench_close: pd.Series) -> pd.Series:
    """
    用收盘价计算主动收益（日频）：pct_change(etf) - pct_change(bench)，并对齐日期。
    """
    df = pd.concat([etf_close, bench_close], axis=1, keys=['etf','bench']).dropna()
    if df.empty or len(df) < 2:
        return pd.Series(dtype=float)
    active = df['etf'].pct_change() - df['bench'].pct_change()
    return active.dropna()

def summarize_metrics_by_cycle_type(
    tk: str,
    dfp: pd.DataFrame,
    bench_close: pd.Series,
    periods_ex: list[tuple[str, str, str]],
    valid_range_one: Tuple[Optional[pd.Timestamp], Optional[pd.Timestamp]],
    cum_div: pd.Series,
    cycle_type: str,
    rf_annual: float
) -> dict:
    """
    对某一标的 tk，在“指定周期类型”（牛市/熊市/震荡）下做汇总，口径与沪深300版本一致：
    - 不对已年化指标做算术平均；
    - 而是拼接该类型下各区间的“原始日度数据”，计算整体的年化收益/波动/夏普/信息比率/跟踪误差等；
    - 最大回撤：同类区间中取“最深回撤”（数值最小）；
    - 分红：同类区间内分红金额求和；分红率按“全样本日均价 + 总交易日数”年化。
    """
    data_start, data_end = valid_range_one
    if data_start is None or data_end is None:
        return {}

    seg_rets: List[pd.Series] = []
    seg_actives: List[pd.Series] = []
    seg_closes: List[pd.Series] = []
    seg_vols: List[pd.Series] = []
    seg_mdds: List[float] = []
    div_total = 0.0
    has_div = (not is_index_ticker(tk)) and (cum_div is not None) and (not cum_div.empty)

    eff_starts: List[pd.Timestamp] = []
    eff_ends: List[pd.Timestamp] = []

    for start_str, end_str, ptype_raw in periods_ex:
        if normalize_market_cycle(ptype_raw) != cycle_type:
            continue

        s0 = month_start(start_str)
        e0 = month_end(end_str)

        # 与标的有效数据交集
        s = max(s0, data_start)
        e = min(e0, data_end)
        if s >= e:
            continue

        seg = dfp.loc[(dfp.index >= s) & (dfp.index <= e)].copy()
        if seg.empty or 'Close' not in seg.columns:
            continue

        seg_close = seg['Close'].astype(float).dropna()
        if seg_close.empty or len(seg_close) < 2:
            continue

        eff_starts.append(seg_close.index.min())
        eff_ends.append(seg_close.index.max())

        # 区间内 pct_change，避免跨段收益
        seg_ret = seg_close.pct_change().dropna()
        if len(seg_ret) > 0:
            seg_rets.append(seg_ret)

        bench_seg_close = bench_close.loc[seg_close.index.min():seg_close.index.max()]
        act = _active_ret_from_close(seg_close, bench_seg_close)
        if len(act) > 0:
            seg_actives.append(act)

        seg_closes.append(seg_close)

        if 'Volume' in seg.columns:
            seg_vols.append(seg['Volume'].astype(float).dropna())

        seg_mdds.append(max_drawdown(seg_close))

        if has_div:
            div_total += float(dividend_in_period(cum_div, s, e))

    if not eff_starts or not eff_ends:
        return {}

    all_ret = pd.concat(seg_rets, axis=0) if seg_rets else pd.Series(dtype=float)
    all_ret = all_ret[~all_ret.index.duplicated(keep='last')].dropna()

    all_active = pd.concat(seg_actives, axis=0) if seg_actives else pd.Series(dtype=float)
    all_active = all_active[~all_active.index.duplicated(keep='last')].dropna()

    all_close = pd.concat(seg_closes, axis=0) if seg_closes else pd.Series(dtype=float)
    all_close = all_close[~all_close.index.duplicated(keep='last')].dropna()

    all_vol = pd.concat(seg_vols, axis=0) if seg_vols else pd.Series(dtype=float)
    all_vol = all_vol[~all_vol.index.duplicated(keep='last')].dropna()

    ann_ret = annualize_return(all_ret) if len(all_ret) else np.nan
    ann_vol = annualize_vol(all_ret) if len(all_ret) else np.nan
    sr = sharpe_ratio(all_ret, rf_annual=rf_annual) if len(all_ret) else np.nan

    if len(all_active) >= 2:
        alpha = all_active.mean() * 252.0
        te_ann = all_active.std(ddof=1) * np.sqrt(252.0)
        ir = np.nan if te_ann == 0 or np.isnan(te_ann) else alpha / te_ann
    else:
        ir = np.nan
        te_ann = np.nan

    # 最大回撤：回撤为负值，取最小值代表“最深回撤”
    mdd = np.nan if len(seg_mdds) == 0 else float(np.nanmin(seg_mdds))

    avg_vol = all_vol.mean() if len(all_vol) else np.nan

    if is_index_ticker(tk) or (not has_div):
        div_amt = np.nan
        div_y_ann = np.nan
    else:
        div_amt = div_total
        if len(all_close) == 0:
            div_y_ann = np.nan
        else:
            avg_price = float(all_close.mean())
            if avg_price <= 0:
                div_y_ann = np.nan
            else:
                div_y_ann = (div_amt / avg_price) * (252.0 / len(all_close))

    start_dt = min(eff_starts)
    end_dt = max(eff_ends)
    start_str_out = start_dt.strftime("%Y-%m")
    end_str_out = end_dt.strftime("%Y-%m")

    return dict(
        start_str=start_str_out,
        end_str=end_str_out,
        period_type=cycle_type,
        ann_ret=ann_ret,
        ann_vol=ann_vol,
        mdd=mdd,
        sr=sr,
        ir=ir,
        avg_vol=avg_vol,
        te_ann=te_ann,
        div_amt=div_amt,
        div_y_ann=div_y_ann
    )
# ======= 新增结束 =======

# ---------- 可视化（按区间出图 + 图例） ----------
def _fmt_period_name(i: int, s: str, e: str) -> str:
    return f"区间{i}_{s}_{e}"

def plot_nav_line_period(ticker: str, seg_close: pd.Series, i: int, s: str, e: str):
    fig_path = DIR_OUT / f"{ticker}_{_fmt_period_name(i,s,e)}_NAV_line.png"
    plt.figure(figsize=(10,5))
    plt.plot(seg_close.index, seg_close.values, linewidth=1.2, label="收盘价")
    plt.title(f"{ticker} 收盘价走势（{_fmt_period_name(i,s,e)}）")
    plt.xlabel("日期"); plt.ylabel("价格")
    plt.legend(loc="best")
    plt.tight_layout(); plt.savefig(fig_path, dpi=150, bbox_inches='tight'); plt.close()

def resample_monthly_ohlc(df_price: pd.DataFrame) -> pd.DataFrame:
    """将日频 OHLC 重采样为月频（每月一根蜡烛），用于均线计算的‘全历史’基表。"""
    ohlc = df_price[['Open','High','Low','Close']].dropna()
    if ohlc.empty:
        return pd.DataFrame()
    ohlc_m = pd.DataFrame({
        'Open' : ohlc['Open'].resample('M').first(),
        'High' : ohlc['High'].resample('M').max(),
        'Low'  : ohlc['Low'].resample('M').min(),
        'Close': ohlc['Close'].resample('M').last(),
    }).dropna()
    return ohlc_m

def plot_monthly_candle_with_ma_period(ticker: str,
                                       monthly_full: pd.DataFrame,
                                       ma_full: dict[str, pd.Series],
                                       i: int, s: str, e: str):
    """先基于全历史算好月均线，再在区间内切片绘图，避免区间起点均线为空。"""
    if monthly_full.empty:
        return
    # 若传入是 'YYYY-MM'，则取月初/月末；否则按普通日期解析
    if isinstance(s, str) and len(s) == 7:
        s_dt = month_start(s)
    else:
        s_dt = pd.to_datetime(s)

    if isinstance(e, str) and len(e) == 7:
        e_dt = month_end(e)
    else:
        e_dt = pd.to_datetime(e)

    # 区间切片（注意月末索引）
    ohlc_m_seg = monthly_full.loc[(monthly_full.index >= s_dt) & (monthly_full.index <= e_dt)]
    if ohlc_m_seg.empty:
        return

    # 先对齐到 ohlc_m_seg.index，再判断是否全 NaN
    ma5_seg  = ma_full['ma5'].reindex(ohlc_m_seg.index)
    ma10_seg = ma_full['ma10'].reindex(ohlc_m_seg.index)
    ma20_seg = ma_full['ma20'].reindex(ohlc_m_seg.index)

    ap = []
    for lab, ser in [('MA5(月)', ma5_seg), ('MA10(月)', ma10_seg), ('MA20(月)', ma20_seg)]:
        # 只要有一个有效值，就可以画；传入“含 NaN 的对齐序列”，这样 x/y 维度一致
        if ser.notna().any():
            ap.append(mpf.make_addplot(ser, panel=0, width=1.0, label=lab))

    # 如果三条均线全是 NaN，则不传 addplot
    ap = ap if ap else None


    s_mpf = mpf.make_mpf_style(
        base_mpf_style='yahoo',
        rc={'font.sans-serif': matplotlib.rcParams['font.sans-serif'],
            'axes.unicode_minus': False}
    )
    fig_path = DIR_OUT / f"{ticker}_区间{i}_{s}_{e}_Monthly_Candle_MA.png"
    mpf.plot(ohlc_m_seg, type='candle', addplot=ap, volume=False, style=s_mpf, ylabel='价格',
             title=f"{ticker} 月K线（含均线，区间{i}_{s}_{e}）",
             figscale=1.1, figratio=(16,9), tight_layout=True,
             savefig=dict(fname=str(fig_path), dpi=150, bbox_inches='tight'),
             warn_too_much_data=1000000, datetime_format='%Y-%b')


# ---------- 表格格式化（中文列名 + 保留位数） ----------
CN_COLUMNS = [
    "标的",               # Ticker
    "开始时间",           # Period Start
    "结束时间",           # Period End
    "市场周期",           # Bull/Bear/Range
    "年化收益率",
    "年化波动率",
    "最大回撤",
    "夏普比率",
    "信息比率",
    "日均成交量",
    "年化跟踪误差",
    "分红",               # 区间内现金分红（元/份）
    "分红率"           # 区间内年化分红率（0.03 表示 3%）
]

def _format_row_cn(ticker: str,
                   start_str: str, end_str: str, period_type: str,
                   ann_ret: float, ann_vol: float, mdd: float,
                   sr: float, ir: float, avg_vol: float,
                   te_ann: float, div_amt: float, div_yield_ann: float) -> dict:
    # 百分比字段以“数值”保存：0.03 表示 3%，到 Excel 再设置百分比格式
    def pct_num(v):
        return None if pd.isna(v) else float(v)
    def two(v):
        return None if pd.isna(v) else round(float(v), 2)
    def num(v):
        return None if pd.isna(v) else float(v)
    def vol_in_million(v):
        return None if pd.isna(v) else int(round(float(v) / 1_000_000.0))
    return {
        "标的": ticker,
        "开始时间": start_str,
        "结束时间": end_str,
        "市场周期": period_type,
        "年化收益率": pct_num(ann_ret),
        "年化波动率": pct_num(ann_vol),
        "最大回撤": pct_num(mdd),
        "夏普比率": two(sr),
        "信息比率": two(ir),
        "日均成交量": vol_in_million(avg_vol),
        "年化跟踪误差": pct_num(te_ann),
        "分红": num(div_amt),
        "分红率": pct_num(div_yield_ann)
    }

# ---------------- 主流程 ----------------
def main():
    # —— 将不同来源统一为 (start,end,type) 三元组 —— #
    periods_ex: list[tuple[str, str, str]] = normalize_periods(
        PERIOD_SOURCE, PERIODS, PERIODS_TABLE_PATH, PERIODS_TABLE_SHEET
    )

    all_prices: Dict[str, pd.DataFrame] = {}
    valid_range: Dict[str, Tuple[Optional[pd.Timestamp], Optional[pd.Timestamp]]] = {}
    # 缓存 ETF 累计分红
    all_cum_div: Dict[str, pd.Series] = {}

    # 1) 下载、保存原始、记录区间
    for tk in TICKERS:
        df = fetch_full_history_by_ticker(tk)
        all_prices[tk] = df
        save_raw_to_excel(tk, df)
        dfx = df.dropna(subset=["Close"])
        valid_range[tk] = (None, None) if dfx.empty else (dfx.index.min(), dfx.index.max())
        # —— 若为 ETF，则同时抓取累计分红 —— #
        if not is_index_ticker(tk):
            try:
                all_cum_div[tk] = fetch_etf_dividend_cumulative(tk)
            except Exception:
                all_cum_div[tk] = pd.Series(dtype=float)

    # 2) brief 写 raw_data
    with open(BRIEF_TXT, "w", encoding="utf-8") as f:
        f.write(f"获取时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        for tk in TICKERS:
            s, e = valid_range[tk]
            f.write(f"{tk}: 有效区间 {'暂无有效收盘价' if (s is None or e is None) else f'{s.date()} — {e.date()}'}\n")

    # 3) 指标计算（逐区间），并**按区间出图**
    # 基准直接使用 INDEX_TICKERS[0]
    bench_close = all_prices[INDEX_TICKERS[0]]['Close'].astype(float)
    all_rows_cn = []

    for tk in TICKERS:
        dfp = all_prices[tk].copy()
        close_full = dfp['Close'].astype(float)
        cum_div = all_cum_div.get(tk, pd.Series(dtype=float))  # 仅 ETF 有

        # —— 关键新增：先基于全历史生成月频与均线 —— #
        monthly_full = resample_monthly_ohlc(dfp)
        ma_full = {
            'ma5' : monthly_full['Close'].rolling(5).mean(),
            'ma10': monthly_full['Close'].rolling(10).mean(),
            'ma20': monthly_full['Close'].rolling(20).mean(),
        }
        # ———————————————————————————————— #

        for i, (start_str, end_str, period_type) in enumerate(periods_ex, 1):
            # start_str/end_str 是 'YYYY-MM'
            s = month_start(start_str)
            e = month_end(end_str)
            seg = dfp.loc[(dfp.index >= s) & (dfp.index <= e)].copy()

            if seg.empty:
                all_rows_cn.append(_format_row_cn(
                    tk, start_str, end_str, period_type,
                    np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan
                ))
                continue

            seg_close = seg['Close'].astype(float)
            seg_ret   = seg_close.pct_change()

            # 对齐同区间的基准（按日期切片）
            bench_seg_close = bench_close.loc[seg.index.min():seg.index.max()]
            bench_seg_ret   = bench_seg_close.pct_change()

            ann_ret = annualize_return(seg_ret)
            ann_vol = annualize_vol(seg_ret)
            mdd     = max_drawdown(seg_close)
            sr      = sharpe_ratio(seg_ret, rf_annual=RISK_FREE_ANNUAL)
            ir      = information_ratio(seg_ret, bench_seg_ret)
            avg_vol = seg.get('Volume', pd.Series(dtype=float)).dropna().mean()

            # 区间年化跟踪误差
            te_ann  = tracking_error_annualized_period(seg_close, bench_seg_close)

            # 区间分红与年化分红率（仅 ETF 有；指数为 NaN）
            if is_index_ticker(tk) or cum_div is None or cum_div.empty:
                div_amt = np.nan
                div_y_ann = np.nan
            else:
                div_amt = dividend_in_period(cum_div, s, e)  # 元/份
                div_y_ann = dividend_yield_annualized(div_amt, seg_close)  # 0.03 表示 3%

            all_rows_cn.append(_format_row_cn(
                tk, start_str, end_str, period_type,
                ann_ret, ann_vol, mdd, sr, ir, avg_vol, te_ann, div_amt, div_y_ann
            ))

            # —— 与区间对应的出图 —— 
            if VISUALIZE:
                plot_nav_line_period(tk, seg_close, i, start_str, end_str)
                plot_monthly_candle_with_ma_period(tk, monthly_full, ma_full, i, start_str, end_str)

    # 4) 写一个汇总 Excel（中文列名，已格式化）
    metrics_df_cn = pd.DataFrame(all_rows_cn, columns=CN_COLUMNS)

    # 排序：先按“开始时间”，再按“标的”
    metrics_df_cn = metrics_df_cn.sort_values(by=["开始时间", "标的"], kind="mergesort")

    # 全空判断
    _keep_cols  = ["标的", "开始时间", "结束时间", "市场周期"]
    _other_cols = [c for c in metrics_df_cn.columns if c not in _keep_cols]

    def _is_empty_cell(x):
        return pd.isna(x) or (isinstance(x, str) and x.strip() == "")

    if _other_cols:
        _all_empty = metrics_df_cn[_other_cols].applymap(_is_empty_cell).all(axis=1)
        metrics_df_cn = metrics_df_cn.loc[~_all_empty].copy()

    # ======= 新增：为每个标的追加“整体区间”汇总行 ======= #
    if not metrics_df_cn.empty:
        # 基于周期配置的整体起止时间
        periods_start_dt = min(month_start(s) for s, _, _ in periods_ex)
        periods_end_dt = max(month_end(e) for _, e, _ in periods_ex)

        agg_rows = []
        for tk in TICKERS:
            data_start, data_end = valid_range.get(tk, (None, None))
            if data_start is None or data_end is None:
                continue

            # 标的有效时间区间 ∩ 周期整体时间区间
            s_dt = max(periods_start_dt, data_start)
            e_dt = min(periods_end_dt, data_end)
            if s_dt >= e_dt:
                continue

            dfp = all_prices[tk]
            seg = dfp.loc[(dfp.index >= s_dt) & (dfp.index <= e_dt)].copy()
            if seg.empty:
                continue

            seg_close = seg['Close'].astype(float)
            seg_ret = seg_close.pct_change()

            bench_seg_close = bench_close.loc[seg.index.min():seg.index.max()]
            bench_seg_ret = bench_seg_close.pct_change()

            ann_ret = annualize_return(seg_ret)
            ann_vol = annualize_vol(seg_ret)
            mdd = max_drawdown(seg_close)
            sr = sharpe_ratio(seg_ret, rf_annual=RISK_FREE_ANNUAL)
            ir = information_ratio(seg_ret, bench_seg_ret)
            avg_vol = seg.get('Volume', pd.Series(dtype=float)).dropna().mean()

            te_ann = tracking_error_annualized_period(seg_close, bench_seg_close)

            cum_div = all_cum_div.get(tk, pd.Series(dtype=float))
            if is_index_ticker(tk) or cum_div is None or cum_div.empty:
                div_amt = np.nan
                div_y_ann = np.nan
            else:
                div_amt = dividend_in_period(cum_div, s_dt, e_dt)
                div_y_ann = dividend_yield_annualized(div_amt, seg_close)

            # 格式化整体区间的开始/结束月份，例如 2005-06 ～ 2025-10
            start_str_full = s_dt.strftime("%Y-%m")
            end_str_full = e_dt.strftime("%Y-%m")

            agg_rows.append(_format_row_cn(
                tk, start_str_full, end_str_full, "整体",
                ann_ret, ann_vol, mdd, sr, ir, avg_vol, te_ann, div_amt, div_y_ann
            ))

        if agg_rows:
            agg_df = pd.DataFrame(agg_rows, columns=CN_COLUMNS)
            # 直接追加在当前表格最后若干行
            metrics_df_cn = pd.concat([metrics_df_cn, agg_df], ignore_index=True)
    # ======= 新增结束 ======= #

    # ======= 新增：按市场周期类型（牛市/熊市/震荡）汇总（口径与沪深300一致） ======= #
    cycle_rows_cn = []
    cycle_types = ["牛市", "熊市", "震荡"]

    for tk in TICKERS:
        dfp = all_prices[tk].copy()
        cum_div = all_cum_div.get(tk, pd.Series(dtype=float))
        vr = valid_range.get(tk, (None, None))

        for ct in cycle_types:
            res = summarize_metrics_by_cycle_type(
                tk=tk,
                dfp=dfp,
                bench_close=bench_close,
                periods_ex=periods_ex,
                valid_range_one=vr,
                cum_div=cum_div,
                cycle_type=ct,
                rf_annual=RISK_FREE_ANNUAL
            )
            if not res:
                cycle_rows_cn.append(_format_row_cn(
                    tk, "", "", ct,
                    np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan
                ))
                continue

            cycle_rows_cn.append(_format_row_cn(
                tk,
                res["start_str"],
                res["end_str"],
                res["period_type"],
                res["ann_ret"],
                res["ann_vol"],
                res["mdd"],
                res["sr"],
                res["ir"],
                res["avg_vol"],
                res["te_ann"],
                res["div_amt"],
                res["div_y_ann"]
            ))

    cycle_df_cn = pd.DataFrame(cycle_rows_cn, columns=CN_COLUMNS)
    if not cycle_df_cn.empty:
        order_map = {"牛市": 0, "熊市": 1, "震荡": 2}
        cycle_df_cn["_order"] = cycle_df_cn["市场周期"].map(lambda x: order_map.get(str(x), 99))
        cycle_df_cn = cycle_df_cn.sort_values(by=["标的", "_order"], kind="mergesort").drop(columns=["_order"])
    # ======= 新增结束 ======= #

    with pd.ExcelWriter(ALL_METRICS_XLSX, engine="openpyxl") as writer:
        metrics_df_cn.to_excel(writer, sheet_name="指标汇总", index=False)

        # ======= 新增：写入“按市场周期分组”sheet ======= #
        cycle_df_cn.to_excel(writer, sheet_name="按市场周期分组", index=False)
        # ======= 新增结束 ======= #

        # 设置 Excel 显示为百分比格式（保存数值为 0.03，显示为 3%）
        ws = writer.sheets["指标汇总"]
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        # 合并单元格：相邻行“开始时间/结束时间/市场周期”均相同则纵向合并
        from openpyxl.styles import Alignment

        _merge_keys = ["开始时间", "结束时间", "市场周期"]
        # 需要合并的三列的 Excel 列号（1-based）
        _merge_cols = {k: header.index(k) + 1 for k in _merge_keys if k in header}

        if len(_merge_cols) == 3 and not metrics_df_cn.empty:
            # 只看这三列，按当前排序逐段合并
            key_df = metrics_df_cn[_merge_keys].astype(str).fillna("")
            n = len(key_df)
            r = 0
            while r < n:
                # 向下找到与当前行完全相同的最长连续段 [r, r2]
                r2 = r
                while r2 + 1 < n and (key_df.iloc[r2 + 1] == key_df.iloc[r]).all():
                    r2 += 1
                # 需要合并且长度>1
                if r2 > r:
                    excel_row_start = r + 2   # DataFrame写入后，数据从第2行开始
                    excel_row_end   = r2 + 2
                    for k, cidx in _merge_cols.items():
                        ws.merge_cells(start_row=excel_row_start, start_column=cidx,
                                       end_row=excel_row_end,   end_column=cidx)
                        # 让合并后的顶端单元格垂直居中
                        ws.cell(row=excel_row_start, column=cidx).alignment = Alignment(vertical="center")
                r = r2 + 1

        # 显示百分比
        pct_cols = {
            "年化收益率",
            "年化波动率",
            "最大回撤",
            "年化跟踪误差",
            "分红率"
        }
        pct_col_idxs = [header.index(col) + 1 for col in header if col in pct_cols]
        from openpyxl.styles import numbers
        for col_idx in pct_col_idxs:
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                for c in cell:
                    c.number_format = '0.00%'  # 两位小数的百分比显示

        # 把 日均成交量 按 M 显示
        if "日均成交量" in header:
            vol_col_idx = header.index("日均成交量") + 1  # openpyxl 列索引从 1 开始
            for cell in ws.iter_cols(min_col=vol_col_idx, max_col=vol_col_idx, min_row=2):
                for c in cell:
                    # 如果该单元格是数字，就按 “百万 + M” 显示；文本则跳过
                    if isinstance(c.value, (int, float)):
                        c.number_format = '#,##0"M"'

        # ======= 新增：对“按市场周期分组”sheet 同样做百分比与成交量格式化 ======= #
        ws2 = writer.sheets["按市场周期分组"]
        header2 = [cell.value for cell in next(ws2.iter_rows(min_row=1, max_row=1))]

        pct_col_idxs2 = [header2.index(col) + 1 for col in header2 if col in pct_cols]
        for col_idx in pct_col_idxs2:
            for cell in ws2.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                for c in cell:
                    c.number_format = '0.00%'

        if "日均成交量" in header2:
            vol_col_idx2 = header2.index("日均成交量") + 1
            for cell in ws2.iter_cols(min_col=vol_col_idx2, max_col=vol_col_idx2, min_row=2):
                for c in cell:
                    if isinstance(c.value, (int, float)):
                        c.number_format = '#,##0"M"'
        # ======= 新增结束 ======= #

if __name__ == "__main__":
    main()
